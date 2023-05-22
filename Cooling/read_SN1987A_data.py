import numpy as np

#Importing loader of excel sheet
from openpyxl import load_workbook

#COLLECTING THE DATA

file_path = 'Data_SN1987A.xlsx'

#Coletando arquivo de dados
wp = load_workbook(file_path)
#Utilizando apenas uma das planilhas do arquivo Excel (chamada 'Dados')
ws = wp.active

#Gerando uma tupla com as colunas da tabela
cols = ws.iter_cols(min_row=3,max_col=17,max_row=19)

#Listas de váriáveis das colunas
tempo_K = []
tempo_I = []
tempo_B = []

E_K = []
iE_K = []
theta_K = []
itheta_K = []

E_I = []
iE_I = []
theta_I = []
itheta_I = []

E_B = []
iE_B = []

#Percorrendo cada coluna e inserindo nas listas
for col in cols:
    for cell in col:
        if cell.column == 2 and cell.value != None:
            tempo_K.append(cell.value)
        elif cell.column == 3 and cell.value != None:
            E_K.append(cell.value)
        elif cell.column == 4 and cell.value != None:
            iE_K.append(cell.value)
        elif cell.column == 5 and cell.value != None:
            theta_K.append(cell.value)
        elif cell.column == 6 and cell.value != None:
            itheta_K.append(cell.value)

        elif cell.column == 9 and cell.value != None:
            tempo_I.append(cell.value)
        elif cell.column == 10 and cell.value != None:
            E_I.append(cell.value)
        elif cell.column == 11 and cell.value != None:
            iE_I.append(cell.value)
        elif cell.column == 12 and cell.value != None:
            theta_I.append(cell.value)
        elif cell.column == 13 and cell.value != None:
            itheta_I.append(cell.value)
        
        elif cell.column == 15 and cell.value != None:
            tempo_B.append(cell.value)
        elif cell.column == 16 and cell.value != None:
            E_B.append(cell.value)
        elif cell.column == 17 and cell.value != None:
            iE_B.append(cell.value)

#Kamiokande-II background from events from 1-16 in Hz/MeV (https://arxiv.org/pdf/astro-ph/0608399.pdf)
# background_KII = [1.0e-5, 5.4e-4, 3.1e-2, 8.5e-3, 5.3e-4, 7.1e-2, 5.0e-6, 1.0e-5, 1.0e-5, 1.8e-2, 4.0e-4, 1.4e-2, 7.3e-2, 5.2e-2, 1.8e-2, 7.3e-2]

#KII background following https://arxiv.org/pdf/1409.4710.pdf (Comparative...)
B_rate_K = [1.0e-5, 5.4e-4, 2.4e-2, 2.8e-3, 5.3e-4, 7.9e-2, 5e-6, 1e-5, 1e-5, 4.2e-3, 4e-4, 3.2e-3, 7.3e-2, 5.3e-2, 1.8e-2, 7.3e-2]
B_rate_B = [8.4e-4, 1.3e-3, 1.2e-3, 1.3e-3, 1.3e-3]

#Maria Laura Costantini, et.al.
# B_rate_K=[1.0*10**-5, 5.4*10**-4, 3.1*10**-2, 8.5*10**-3, 5.3*10**-4, 7.1*10**-2, 5.0*10**-6, 1.0*10**-5, 1.0*10**-5, 1.8*10**-2, 4.0*10**-4, 1.4*10**-2, 7.3*10**-2, 5.2*10**-2, 1.8*10**-2, 7.3*10**-2]
#Lamb and Loredo
#B_rate_K=[1.6*10**-5, 1.9*10**-3, 2.9*10**-2, 1.2*10**-2, 2.1*10**-3, 3.7*10**-2, 4.5*10**-5, 8.2*10**-5, 1.5*10**-5, 1.5*10**-2, 1.9*10**-3, 1.6*10**-2, 3.8*10**-2, 2.9*10**-2, 2.8*10**-2, 3.8*10**-2]
#Baksan
# B_rate_B=[8.4*10**-4, 1.3*10**-3, 1.2*10**-3, 1.3*10**-3, 1.3*10**-3]

#print('Delta E =', max(E_B+E_I+E_K) - min(E_B+E_I+E_K))
#print('Delta T =', max(tempo_B+tempo_I+tempo_K) - min(tempo_B+tempo_I+tempo_K))
#print('mean E =', np.mean(E_B+E_I+E_K))
#print(min(tempo_K+tempo_B+tempo_I))
#print(min(E_K+E_I+E_B))