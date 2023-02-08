import numpy as np

#Importing loader of excel sheet
from openpyxl import load_workbook

#COLLECTING THE DATA

file_path = 'Data_SN1987A.xlsx'

#Coletando arquivo de dados
wp = load_workbook(file_path)
#Utilizando apenas uma das planilhas do arquivo Excel (chamada 'Dados')
ws = wp.active

#Gerando uma tupla com as colunas da tabela
cols = ws.iter_cols(min_row=3,max_col=17,max_row=19)

#Listas de váriáveis das colunas
tempo_K = []
tempo_I = []
tempo_B = []

E_K = []
iE_K = []
theta_K = []
itheta_K = []

E_I = []
iE_I = []
theta_I = []
itheta_I = []

E_B = []
iE_B = []

#Percorrendo cada coluna e inserindo nas listas
for col in cols:
    for cell in col:
        if cell.column == 2 and cell.value != None:
            tempo_K.append(cell.value)
        elif cell.column == 3 and cell.value != None:
            E_K.append(cell.value)
        elif cell.column == 4 and cell.value != None:
            iE_K.append(cell.value)
        elif cell.column == 5 and cell.value != None:
            theta_K.append(cell.value)
        elif cell.column == 6 and cell.value != None:
            itheta_K.append(cell.value)

        elif cell.column == 9 and cell.value != None:
            tempo_I.append(cell.value)
        elif cell.column == 10 and cell.value != None:
            E_I.append(cell.value)
        elif cell.column == 11 and cell.value != None:
            iE_I.append(cell.value)
        elif cell.column == 12 and cell.value != None:
            theta_I.append(cell.value)
        elif cell.column == 13 and cell.value != None:
            itheta_I.append(cell.value)
        
        elif cell.column == 15 and cell.value != None:
            tempo_B.append(cell.value)
        elif cell.column == 16 and cell.value != None:
            E_B.append(cell.value)
        elif cell.column == 17 and cell.value != None:
            iE_B.append(cell.value)

#Kamiokande-II background from events from 1-16 in Hz/MeV (https://arxiv.org/pdf/astro-ph/0608399.pdf)
background_KII = [1.0e-5, 5.4e-4, 3.1e-2, 8.5e-3, 5.3e-4, 7.1e-2, 5.0e-6, 1.0e-5, 1.0e-5, 1.8e-2, 4.0e-4, 1.4e-2, 7.3e-2, 5.2e-2, 1.8e-2, 7.3e-2]

#print('Delta E =', max(E_B+E_I+E_K) - min(E_B+E_I+E_K))
#print('Delta T =', max(tempo_B+tempo_I+tempo_K) - min(tempo_B+tempo_I+tempo_K))
#print('mean E =', np.mean(E_B+E_I+E_K))
#print(min(tempo_K+tempo_B+tempo_I))
#print(min(E_K+E_I+E_B))