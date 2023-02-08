#Importing loader of excel sheet
from openpyxl import load_workbook
import csv

def read_SN():

	#Implementado por: Marcos
	#Uma vez que o arquivo está no seu Drive é provável que o caminho seja
	#file_path = "/home/pedro/Downloads/SN1987A/Dados SN1987A.xlsx"
	file_path = "Dados SN1987A.xlsx"
	#Coletando arquivo de dados
	wp = load_workbook(file_path)
	#Utilizando apenas uma das planilhas do arquivo Excel (chamada 'Dados')
	ws = wp.active
	Dados = wp['Plan1']

	#Obs: pelos testes que fiz, somente o próprio usuário terá acesso ao conteúdo do respectivo Google Drive pessoal
	#(Dedin)- Sim, cada um tem que acessar o próprio drive. Mas da para colocar na pasta compartilhada, aí cada um acessa essa mesma pasta 
	#quando forem rodar o código.

	#Gerando uma tupla com as colunas da tabela
	cols = ws.iter_cols(min_row=3,max_col=17,max_row=19)

	#Listas de váriáveis das colunas
	tempo_K = []
	tempo_I = []
	tempo_B = []

	E_K = []
	iE_K = []
	theta_K = []
	itheta_K = []

	E_I = []
	iE_I = []
	theta_I = []
	itheta_I = []

	E_B = []
	iE_B = []

	#Percorrendo cada coluna e inserindo nas listas
	for col in cols:
		for cell in col:
			if cell.column == 2 and cell.value != None:
				tempo_K.append(cell.value)
			elif cell.column == 3 and cell.value != None:
		    		E_K.append(cell.value)
			elif cell.column == 4 and cell.value != None:
			    iE_K.append(cell.value)
			elif cell.column == 5 and cell.value != None:
			    theta_K.append(cell.value)
			elif cell.column == 6 and cell.value != None:
			    itheta_K.append(cell.value)

			elif cell.column == 7 and cell.value != None:
			    tempo_I.append(cell.value)
			elif cell.column == 8 and cell.value != None:
			    E_I.append(cell.value)
			elif cell.column == 9 and cell.value != None:
			    iE_I.append(cell.value)
			elif cell.column == 10 and cell.value != None:
			    theta_I.append(cell.value)
			elif cell.column == 11 and cell.value != None:
			    itheta_I.append(cell.value)

			elif cell.column == 12 and cell.value != None:
			    tempo_B.append(cell.value)
			elif cell.column == 13 and cell.value != None:
			    E_B.append(cell.value)
			elif cell.column == 14 and cell.value != None:
			    iE_B.append(cell.value)
	return tempo_K,tempo_I,tempo_B, E_K,iE_K,theta_K,itheta_K,E_I,iE_I,theta_I,itheta_I,E_B,iE_B
	

def read_SN_csv():
	file_path = "Dados SN1987A.csv"
	
	tempo_K = []
	tempo_I = []
	tempo_B = []

	E_K = []
	iE_K = []
	theta_K = []
	itheta_K = []

	E_I = []
	iE_I = []
	theta_I = []
	itheta_I = []

	E_B = []
	iE_B = []

	with open(file_path) as csv_file:
		csv_reader = csv.reader(csv_file, delimiter=',')
		line_count = 0
		for row in csv_reader:
			if line_count == 0:
				#print(row[1],row[2],row[3],row[4],row[5],row[8],row[9],row[10],row[11],row[12],row[14],row[15],row[16])
				line_count+=1
			else:
				if row[1] != '':
					tempo_K.append(float(row[1]))
				if row[2] != '':
					E_K.append(float(row[2]))
				if row[3] != '':
					iE_K.append(float(row[3]))
				if row[4] != '':
					 theta_K.append(float(row[4]))
				if row[5] != '':
					itheta_K.append(float(row[5]))
				if row[8] != '':
					tempo_I.append(float(row[8]))
				if row[9] != '':
					E_I.append(float(row[9]))
				if row[10] != '':
					iE_I.append(float(row[10]))
				if row[11] != '':
					theta_I.append(float(row[11]))
				if row[12] != '':
					 itheta_I.append(float(row[12]))
				if row[14] != '':
					tempo_B.append(float(row[14]))
				if row[15] != '':
					E_B.append(float(row[15]))
				if row[16] != '':
					iE_B.append(float(row[16]))
				line_count+=1
	#print(f'Processed {line_count} lines.')
	#print(tempo_K,tempo_I,tempo_B, E_K,iE_K,theta_K,itheta_K,E_I,iE_I,theta_I,itheta_I,E_B,iE_B)
	return tempo_K,tempo_I,tempo_B, E_K,iE_K,theta_K,itheta_K,E_I,iE_I,theta_I,itheta_I,E_B,iE_B
    		
